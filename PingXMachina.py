import tkinter as tk
import os
import random
from tkinter import *
from tkinter.filedialog import askopenfilename
import datetime
import openpyxl
import pygubu
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell

#GUI
root= tk.Tk()
root.title('PingXMachina v1')

canvas1 = tk.Canvas(root, width = 300, height = 300)
canvas1.pack()

labelx = tk.Label(root, text= 'Select a list of machines/ip/ostnames by clicking \n-Run the Machine- \n , .txt file', fg='green', font=('helvetica', 8, 'bold'))
canvas1.create_window(150, 50, window=labelx)

#Ping and write Results to xlsx file.
def pingMachine():
    today=(datetime.datetime.now())

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'servernames'

    redFill = PatternFill(start_color='FF5733',
                       end_color='FF5733',
                       fill_type='solid')

    greenFill = PatternFill(start_color='33FF57',
                       end_color='33FF57',
                       fill_type='solid')

    with open(hello.filenameX, "r") as ins:
        array = []
        count = 1
        for servername in ins:
            array.append(servername)
            response = os.system("ping -n 1 "+ servername)
            if response == 0:
                ws['A'+ str(count)]=(servername)
                ws['B'+ str(count)]=("online")
                ws['B' + str(count)].fill = greenFill
                ws['C' + str(count)]=today.strftime("%H:%M:%S")
                ws['D'+ str(count)]=today.strftime("%m/%d/%Y")
                count+=1
            else:
                ws['A'+ str(count)]=(servername)
                ws['B'+ str(count)]=("offline")
                ws['B' + str(count)].fill = redFill
                ws['C' + str(count)]=today.strftime("%H:%M:%S")
                ws['D'+ str(count)]=today.strftime("%m/%d/%Y")
                count+=1

    ws.column_dimensions["A"].width = 25.0
    wb.save("PingResults.xlsx")

#Button
def hello():  
    Tk().withdraw() # we don't want a full GUI, so keep the root window from appearing
    hello.filenameX = askopenfilename() # show an "Open" dialog box and return the path to the selected file
    print(hello.filenameX)
    pingMachine()

    #Label after its all done.
    label1 = tk.Label(root, text= 'Done! Check Results.xlsx', fg='green', font=('helvetica', 12, 'bold'))
    canvas1.create_window(150, 200, window=label1)

    #Current Working Directory
    cwd = os.getcwd()
    print (cwd)

    os.startfile('PingResults.xlsx')
    
button1 = tk.Button(text='Run the Machine',command=hello, bg='brown',fg='white')
canvas1.create_window(150, 150, window=button1)

root.mainloop()
